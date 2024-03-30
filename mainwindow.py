import os
import openpyxl
from PyQt5.QtWidgets import (
    QMainWindow,
    QFileDialog,
    QHeaderView,
    QMessageBox,
    QApplication,
)
from PyQt5.QtCore import Qt, pyqtSignal, QThread
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from App2 import Ui_MainWindow
from Dialog.Filter.FilterDialog import FilterDialog
from Dialog.Renaming.RenameWindow import RenameDialog
from Dialog.Progress.LoadingScreen import LoadingScreen
from Dialog.Test.TestWindow import TestDialog
from Dialog.Select.SelectWindow import TypeDialog
from pathlib import Path
from helpers.valid import is_valid_name, is_valid_path

class LoadFilesThread(QThread):
    progress_updated = pyqtSignal(int)  # Signal to update the progress bar
    loading_completed = pyqtSignal()

    def __init__(self, folder_path):
        super(LoadFilesThread, self).__init__()
        self.folder_path = folder_path

    def run(self):
        total_files = sum(len(files) for _, _, files in os.walk(self.folder_path))
        loaded_files = 0

        for root, dirs, files in os.walk(self.folder_path):
            for file_name in files:
                file_path = os.path.normpath(os.path.join(root, file_name))

                # Extract file name and extension
                file_name, file_extension = os.path.splitext(file_name)

                # Create QStandardItems for file name and file path
                file_name_item = QStandardItem(file_name + file_extension)
                file_path_item = QStandardItem(file_path)

                # Set data for file path in UserRole + 1
                file_path_item.setData(file_path, Qt.UserRole + 1)

                # Append the items to the model
                self.model.appendRow([file_path_item, file_name_item])

                # Update progress
                loaded_files += 1
                percent = loaded_files / total_files * 100
                self.progress_updated.emit(int(percent))

        self.loading_completed.emit()

class MainWindow(QMainWindow):
    def __init__(self, logged_username):
        super(MainWindow, self).__init__()

        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.loading = LoadingScreen(self)
        self.selectDialog = TypeDialog(self)
        # Init Elements
        self.open_btn = self.ui.open_btn
        # self.save_btn = self.ui.save_btn
        self.export_btn = self.ui.export_btn
        self.import_btn = self.ui.import_btn
        self.table_view = self.ui.tableView
        self.username = self.ui.name
        self.rename_btn = self.ui.rename_btn
        self.test_btn = self.ui.test_btn
        self.filter_btn = self.ui.filter_btn
        # Connect signals and slots
        self.open_btn.clicked.connect(self.show_type_dialog)
        self.test_btn.clicked.connect(self.show_test_dialog)
        # self.save_btn.clicked.connect(self.rename_folders_excel)
        self.rename_btn.clicked.connect(self.show_rename_dialog)
        self.export_btn.clicked.connect(self.export_to_xlsx)
        self.import_btn.clicked.connect(self.import_xlsx)
        self.filter_btn.clicked.connect(self.show_filter_dialog)
        # Create a standard item model for the table view
        self.model = QStandardItemModel()
        self.table_view.setModel(self.model)
        self.username.setText(logged_username)
        self.folder_signal_connected = False
        self.file_signal_connected = False
        # Status
    def show_test_dialog(self):
        selected_path = QFileDialog.getExistingDirectory(None, "Select Directory", os.path.expanduser('~'))
        # Ensure a path was selected
        if selected_path:
            self.create_test(selected_path)

    def create_test(self, root_path):

        self.child_dirs = ["TestChild1", "TestChild2", "TestChild3", "TestChild4"]
        self.grandchild_dirs = ["TestChild5"]
        self.files = ["testfile.txt", "testfile1.txt", "testfile2.txt", "testfile3.txt", "testfile4.txt", "testfile5.txt"]

        # Create root directory
        root_dir = os.path.join(root_path, "TestParent")
        os.makedirs(root_dir, exist_ok=True)

        # Create child directories
        for child_dir in self.child_dirs:
            child_path = os.path.join(root_dir, child_dir)
            os.makedirs(child_path, exist_ok=True)

            # Create grandchild directories
            if child_dir == "TestChild1":
                for grandchild_dir in self.grandchild_dirs:
                    grandchild_path = os.path.join(child_path, grandchild_dir)
                    os.makedirs(grandchild_path, exist_ok=True)

                    # Create files in grandchild directory
                    for file in self.files[:2]:
                        open(os.path.join(grandchild_path, file), 'a').close()

            # Create files in child directories
            else:
                for file in self.files[2:]:
                    open(os.path.join(child_path, file), 'a').close()

        dialog = TestDialog(self)
        dialog.exec_()

    def show_rename_dialog(self):
        if self.model.rowCount() == 0:
            # Show a warning message
            QMessageBox.warning(self, "Warning", "No data in the table to rename.")
            return
        dialog = RenameDialog(self)
        dialog.confirm_signal.connect(self.rename_confirmation)
        dialog.custom_signal.connect(self.rename_folders_excel)
        dialog.exec_()

    def show_filter_dialog(self):
        dialog = FilterDialog(self)
        dialog.exec_()
    def rename_confirmation(self, new_name):
        if self.is_file_path():
            self.rename_files(new_name)
        else:
            self.rename_folders(new_name)

    def is_file_path(self):
        item = self.model.item(0, 0)
        item_path = item.data(
            Qt.UserRole + 1
        )  # Assuming UserRole + 1 is used to store the path
        if item_path is not None:
            return os.path.isfile(item_path)

    def show_file_dialog(self):
        folder_path = QFileDialog.getExistingDirectory(self, "Select Directory")
        if folder_path:
            self.load_files(folder_path)
        self.selectDialog.hide()

    def show_folder_dialog(self):
        folder_path = QFileDialog.getExistingDirectory(self, "Select Directory")
        if folder_path:
            self.load_folder_contents(folder_path)
        self.selectDialog.hide()

    def show_type_dialog(self):
        # Disconnect signals if they are connected
        if self.folder_signal_connected:
            self.selectDialog.folder_signal.disconnect(self.show_folder_dialog)
            self.folder_signal_connected = False

        if self.file_signal_connected:
            self.selectDialog.file_signal.disconnect(self.show_file_dialog)
            self.file_signal_connected = False

        # Connect signals to slots if they are not connected
        if not self.folder_signal_connected:
            self.selectDialog.folder_signal.connect(self.show_folder_dialog)
            self.folder_signal_connected = True

        if not self.file_signal_connected:
            self.selectDialog.file_signal.connect(self.show_file_dialog)
            self.file_signal_connected = True

        # Show the dialog
        self.selectDialog.exec_()

    def load_folder_contents(self, folder_path):
        # Clear existing items from the model
        self.model.clear()

        # Set up table headers
        self.model.setHorizontalHeaderLabels(["Path", "Name", "New Name"])

        # Set header width to 33% for each column
        header = self.table_view.horizontalHeader()
        header.setSectionResizeMode(
            0, QHeaderView.Stretch
        )  # Path column takes the remaining space
        header.setSectionResizeMode(
            1, QHeaderView.ResizeToContents
        )  # Name column adjusts to content
        header.setSectionResizeMode(
            2, QHeaderView.ResizeToContents
        )  # New Name column adjusts to content
        # Recursively iterate through the folder contents
        self.add_folders_to_model(folder_path)

    def add_folders_to_model(self, folder_path):
        for root, dirs, files in os.walk(folder_path):
            for name in dirs:
                item_path = os.path.normpath(os.path.join(root, name))

                # Create QStandardItems for file path and file name
                folder_path_item = QStandardItem(item_path)
                folder_name_item = QStandardItem(name)

                # Set data for file path in UserRole + 1
                folder_path_item.setData(item_path, Qt.UserRole + 1)

                # Append the items to the model
                self.model.appendRow([folder_path_item, folder_name_item])

    def print_table_data(self):
        for row in range(self.model.rowCount()):
            for col in range(self.model.columnCount()):
                item = self.model.item(row, col)
                if item is not None:
                    item_data = item.data()
                    print(f"Row {row}, Column {col}: {item_data}")
                else:
                    print(f"Row {row}, Column {col}: None")

    def load_files(self, folder_path):
        self.model.clear()
        # Set up table headers
        self.model.setHorizontalHeaderLabels(["Path", "Name", "New Name"])

        # Set header width to 50% for each column
        header = self.table_view.horizontalHeader()
        header.setSectionResizeMode(
            0, QHeaderView.Stretch
        )  # File Name column takes the remaining space
        header.setSectionResizeMode(
            1, QHeaderView.ResizeToContents
        )  # File Path column takes the remaining space
        header.setSectionResizeMode(
            2, QHeaderView.ResizeToContents
        )  # New Name column adjusts to content

        # Create and start the thread
        self.load_files_thread = LoadFilesThread(folder_path)
        self.load_files_thread.model = self.model
        self.loading.show()
        self.load_files_thread.progress_updated.connect(
            self.loading.loadingBar.setValue
        )
        self.load_files_thread.loading_completed.connect(self.loading_completed)
        self.load_files_thread.start()

        # Add files to the model
        # self.add_files_to_model(folder_path)

    def loading_completed(self):
        self.loading.loadingBar.setValue(0)
        self.loading.hide()

    def rename_folders(self, new_name):
        try:
            items = self.get_sorted_items_by_depth()
            index = 1  # Initialize an index to make the new name unique
            for row, item in enumerate(items):  # Assuming you iterate over rows somehow
                old_path = item.data(Qt.UserRole + 1)
                new_path = os.path.join(os.path.dirname(old_path), f"{new_name}_{index}")
                # Check if the new path already exists, and increment the index if needed
                while os.path.exists(new_path):
                    index += 1
                    new_path = os.path.join(
                        os.path.dirname(old_path), f"{new_name}_{index}"
                    )
                os.rename(old_path, new_path)
                item.setData(new_path, Qt.UserRole + 1)

            # Reverse the list before setting the text for items
            items.reverse()
            for row, item in enumerate(items):  # Assuming you iterate over rows somehow
                path_item = self.model.item(row, 0)  # Assuming third column is at index 2
                name_item = self.model.item(row, 1)  # Assuming third column is at index 2
                name = os.path.basename(item.data(Qt.UserRole + 1))
                path_item.setText(item.data(Qt.UserRole + 1))
                name_item.setText(name)
                
        except Exception as e:
                    # Display a warning box with the exception message
                    QMessageBox.warning(self, "Error", str(e))          

    def rename_files(self, new_name):
        try:
            items = self.get_sorted_items_by_depth()  # Assuming this function returns a list of items
            index = 1  # Initialize an index to make the new name unique
            for row, item in enumerate(items):  # Assuming you iterate over rows somehow
                old_path = item.data(Qt.UserRole + 1)
                if os.path.isfile(old_path):  # Check if it's a file
                    file_name, file_ext = os.path.splitext(os.path.basename(old_path))
                    new_path = os.path.join(os.path.dirname(old_path), f"{new_name}_{index}{file_ext}")
                    # Check if the new path already exists, and increment the index if needed
                    while os.path.exists(new_path):
                        index += 1
                        new_path = os.path.join(
                            os.path.dirname(old_path), f"{new_name}_{index}{file_ext}"
                        )
                    os.rename(old_path, new_path)
                    item.setData(new_path, Qt.UserRole + 1)
                    index += 1  # Increment index for the next file

            # Reverse the list before setting the text for items
            items.reverse()
            for row, item in enumerate(items):  # Assuming you iterate over rows somehow
                path_item = self.model.item(row, 0)  # Assuming third column is at index 2
                name_item = self.model.item(row, 1)  # Assuming third column is at index 2
                name = os.path.basename(item.data(Qt.UserRole + 1))
                path_item.setText(item.data(Qt.UserRole + 1))
                name_item.setText(name)
        except Exception as e:
                    # Display a warning box with the exception message
                    QMessageBox.warning(self, "Error", str(e))
            
    def rename_folders_excel(self):
        error_log = []  # Create an empty list to store errors

        if all(
            self.model.item(row, 2) is None or self.model.item(row, 2).text() == ""
            for row in range(self.model.rowCount())
        ):
            # Show a warning if no values in the "New Name" column
            QMessageBox.warning(self, "Warning", "No 'New Name' values provided.")
            return

        # Sort items by depth in descending order
        items = self.get_sorted_items_by_depth()

        for row in range(len(items)):
            path_item = self.model.item(row, 0)
            name_item = self.model.item(row, 1)
            new_name_item = self.model.item(row, 2)

            old_path = (
                path_item.data(Qt.UserRole + 1)
                if path_item and path_item.data(Qt.UserRole + 1)
                else None
            )

            # Check if the "New Name" item exists and has a value
            if new_name_item and new_name_item.text() != "" and old_path:
                new_name = new_name_item.text()

                # Check if the old path exists
                if os.path.exists(old_path):
                    # Check if the new path already exists
                    new_path = os.path.join(os.path.dirname(old_path), new_name)
                    if os.path.exists(new_path):
                        error_log.append(
                            {
                                "path": old_path,
                                "error": f"Skipping rename for path '{old_path}' as '{new_name}' already exists",
                            }
                        )
                    else:
                        try:
                            # Rename the folder
                            os.rename(old_path, new_path)

                            # Update the model data
                            path_item.setData(new_path, Qt.UserRole + 1)
                            path_item.setText(new_path)
                            name_item.setText(new_name)
                        except FileExistsError:
                            error_log.append(
                                {
                                    "path": old_path,
                                    "error": f"Error renaming '{old_path}' to '{new_path}': File already exists",
                                }
                            )
                else:
                    error_log.append(
                        {
                            "path": old_path,
                            "error": f"Folder does not exist: {old_path}.",
                        }
                    )
            else:
                error_log.append(
                    {
                        "path": old_path,
                        "error": f"Skipping rename for path '{old_path}' as 'New Name' is not provided",
                    }
                )

        # Export error log to Excel
        self.export_error_log_to_excel(error_log)

    def rename_files_excel(self):
        error_log = []  # Create an empty list to store errors

        if all(
            self.model.item(row, 2) is None or self.model.item(row, 2).text() == ""
            for row in range(self.model.rowCount())
        ):
            # Show a warning if no values in the "New Name" column
            QMessageBox.warning(self, "Warning", "No 'New Name' values provided.")
            return

        # Sort items by depth in descending order
        items = self.get_sorted_items_by_depth()

        for row in range(len(items)):
            path_item = self.model.item(row, 0)
            name_item = self.model.item(row, 1)
            new_name_item = self.model.item(row, 2)

            old_path = (
                path_item.data(Qt.UserRole + 1)
                if path_item and path_item.data(Qt.UserRole + 1)
                else None
            )

            # Check if the "New Name" item exists and has a value
            if new_name_item and new_name_item.text() != "" and old_path:
                new_name = new_name_item.text()

                # Check if the old path exists
                if os.path.exists(old_path):
                    # Preserve the file extension
                    old_name, old_extension = os.path.splitext(os.path.basename(old_path))
                    new_name, new_extension = os.path.splitext(new_name)

                    # If the new name does not have an extension, preserve the old extension
                    if not new_extension:
                        new_extension = old_extension

                    # Form the new path with preserved extension
                    new_path = os.path.join(os.path.dirname(old_path), new_name + new_extension)

                    # Check if the new path already exists
                    if os.path.exists(new_path):
                        error_log.append(
                            {
                                "path": old_path,
                                "error": f"Skipping rename for path '{old_path}' as '{new_name}' already exists",
                            }
                        )
                    else:
                        try:
                            # Rename the file
                            os.rename(old_path, new_path)

                            # Update the model data
                            path_item.setData(new_path, Qt.UserRole + 1)
                            path_item.setText(new_path)
                            name_item.setText(new_name + new_extension)
                        except FileExistsError:
                            error_log.append(
                                {
                                    "path": old_path,
                                    "error": f"Error renaming '{old_path}' to '{new_path}': File already exists",
                                }
                            )
                else:
                    error_log.append(
                        {
                            "path": old_path,
                            "error": f"File does not exist: {old_path}.",
                        }
                    )
            else:
                error_log.append(
                    {
                        "path": old_path,
                        "error": f"Skipping rename for path '{old_path}' as 'New Name' is not provided",
                    }
                )

        # Export error log to Excel
        self.export_error_log_to_excel(error_log)
        
    def export_error_log_to_excel(self, error_log):
        if not error_log:
            print("No errors to export.")
            return

        desktop_path = str(Path.home() / "Desktop")
        file_path = os.path.join(desktop_path, "error_log.xlsx")

        # Create a new Excel workbook and select the active sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write headers to the Excel file
        headers = ["path", "error"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Write data to the Excel file
        for row_num, error_entry in enumerate(error_log, 2):
            sheet.cell(row=row_num, column=1, value=error_entry["path"])
            sheet.cell(row=row_num, column=2, value=error_entry["error"])

        # Save the Excel file
        workbook.save(file_path)
        print(f"Error log saved to: {file_path}")

    def get_sorted_items_by_depth(self):
        items = []
        for row in range(self.model.rowCount()):
            item = self.model.item(row, 0)
            items.append(item)

        # Modify the sorting key function to handle None values
        items.sort(
            key=lambda x: (
                x.data(Qt.UserRole + 1).count(os.sep)
                if x and x.data(Qt.UserRole + 1)
                else 0
            ),
            reverse=True,
        )

        return items

    def export_to_xlsx(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", "", "Excel Files (*.xlsx)"
        )

        if file_path:
            # Create a new Excel workbook and select the active sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Write headers to the Excel file
            headers = ["Path", "Name", "New Name"]

            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

            # Write data to the Excel file
            for row in range(self.model.rowCount()):
                name_item = self.model.item(row, 0)
                path_item = self.model.item(row, 1)

                name_value = name_item.text()
                path_value = path_item.text()
                new_value = ""

                sheet.append([name_value, path_value, new_value])

            # Save the Excel file
            workbook.save(file_path)

    def import_xlsx(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open Excel File", "", "Excel Files (*.xlsx)"
        )
        if file_path:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            self.model.clear()

            if headers:
                self.model.setHorizontalHeaderLabels(headers)

                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    path_value = row[0]
                    name_value = row[1]

                    if not name_value or not is_valid_name(name_value):  # Check if name is not empty and valid
                        print(f"Row {row_num}: Invalid name: {name_value}")
                        continue  # Skip this row if name is invalid

                    new_name_value = row[2] if row[2] is not None else ""

                    # Check if path is valid
                    if not is_valid_path(path_value, name_value):
                        print(f"Row {row_num}: Path and name do not match: {path_value}, {name_value}")
                        continue  # Skip this row if path and name do not match

                    # Check if path is valid
                    if os.path.normpath(path_value) != path_value:
                        print(f"Row {row_num}: Invalid path: {path_value}")
                        continue  # Skip this row if path is invalid    

                    # Set the data for the new item
                    path_item = QStandardItem(str(row[0]))
                    path_item.setData(
                        str(row[0]), Qt.UserRole + 1
                    )  # Set the path in UserRole+1

                    name_item = QStandardItem(str(name_value))
                    row_items = [
                        path_item,  # "Path" column
                        name_item,  # "Name" column
                        QStandardItem(str(new_name_value)),  # "New Name" column
                    ]
                    self.model.appendRow(row_items)

                # Set header width to 33% for each column
                header = self.table_view.horizontalHeader()
                header.setSectionResizeMode(
                    0, QHeaderView.Stretch
                )  # Path column takes the remaining space
                header.setSectionResizeMode(
                    1, QHeaderView.ResizeToContents
                )  # Name column adjusts to content
                header.setSectionResizeMode(
                    2, QHeaderView.ResizeToContents
                )  # New Name column adjusts to content

                workbook.close()



